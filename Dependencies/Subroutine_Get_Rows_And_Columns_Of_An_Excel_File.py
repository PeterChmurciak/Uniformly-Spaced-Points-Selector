# -*- coding: utf-8 -*-
"""
Function returning the number of rows and columns of Excel file.
Copyright (C) 2022  Peter Chmurčiak

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see https://www.gnu.org/licenses/.
"""
import openpyxl  # Working with Excel files


def Get_Rows_And_Columns_Of_An_Excel_File(path_to_excel_file):
    # Open Excel file with intention to read only
    excel_workbook_handle = openpyxl.load_workbook(path_to_excel_file)
    # Get sheet names and select the first one (assuming that it is the correct one)
    sheet_names_list = excel_workbook_handle.sheetnames
    first_sheet_name = sheet_names_list[0]
    first_sheet_handle = excel_workbook_handle[first_sheet_name]
    # Extract the sheet properties
    number_of_rows = first_sheet_handle.max_row
    number_of_columns = first_sheet_handle.max_column
    # Return tuple of rows and columns to allow unpacking
    return (number_of_rows, number_of_columns)
