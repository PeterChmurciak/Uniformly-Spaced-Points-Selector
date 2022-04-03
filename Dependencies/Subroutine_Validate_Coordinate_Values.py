# -*- coding: utf-8 -*-
"""
Function validating values of latitude and longitude stored as columns in Excel file.
Copyright (C) 2022  Peter Chmurčiak

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see https://www.gnu.org/licenses/.
"""
import openpyxl  # Working with Excel files


def Validate_Coordinate_Values(path_to_excel_file):
    # Variable for storing the result
    are_they_valid = True
    # Open excel file with intention to only read
    excel_workbook_handle = openpyxl.load_workbook(path_to_excel_file)
    # Get sheet names and select the first one (assuming that it is the correct one)
    sheet_names_list = excel_workbook_handle.sheetnames
    first_sheet_name = sheet_names_list[0]
    first_sheet_handle = excel_workbook_handle[first_sheet_name]
    # Extract the sheet properties
    number_of_rows = first_sheet_handle.max_row
    # Check if latitude and longitude values are witin their bounds - if they acutally are lat/long numeric values
    latitude_column = 2
    longitude_column = 3
    # Go through all lines
    for row in range(2, number_of_rows + 1):
        latitude_reading = first_sheet_handle.cell(row, latitude_column).value
        longitude_reading = first_sheet_handle.cell(row, longitude_column).value
        try:
            latitude_reading = float(latitude_reading)
            longitude_reading = float(longitude_reading)
        except ValueError:
            are_they_valid = False
            break
        else:
            if (latitude_reading < -90 or latitude_reading > 90) or (
                longitude_reading < -180 or longitude_reading > 180
            ):
                are_they_valid = False
                break
    # Return result
    return are_they_valid
