# -*- coding: utf-8 -*-
"""
Function for selecting and sorting uniformly spaced subset of geographical points.
Copyright (C) 2022  Peter Chmurčiak

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see https://www.gnu.org/licenses/.
"""
import os  # Working with file paths
import openpyxl  # Working with Excel files
import openpyxl.styles  # Excel cell formatting
from openpyxl.worksheet.table import Table  # Excel table management
from openpyxl.styles.borders import Border, Side  # Excel cell border formatting
import geopy.point  # Definition of point object with lat/lon coordinates
import geopy.distance  # Distance between point objects in m or km
import numpy  # Statistical functions std,argmin
import webbrowser  # Working with web browser
import gmplot  # Plotting the coordinates using gmaps
import shutil  # Making a backup copy of a file


def Select_The_Points(path_to_excel_file, how_many_points_to_find):

    # Make a copy of the original file with the intention to modify only the copy
    chosen_file_folder_path_and_file_name = os.path.split(path_to_excel_file)
    chosen_file_folder_path = chosen_file_folder_path_and_file_name[0]
    chosen_file_name_and_extension = os.path.splitext(
        chosen_file_folder_path_and_file_name[1]
    )
    copied_file_name = (
        chosen_file_name_and_extension[0]
        + "_MARKED"
        + chosen_file_name_and_extension[1]
    )
    path_to_copied_file = os.path.join(chosen_file_folder_path, copied_file_name)
    shutil.copy(path_to_excel_file, path_to_copied_file)

    # Open the copy of the original file
    excel_workbook_handle = openpyxl.load_workbook(path_to_copied_file)

    # Get sheet names and select the first one (assuming that it is the correct one)
    sheet_names = excel_workbook_handle.sheetnames
    first_sheet_name = sheet_names[0]
    first_sheet = excel_workbook_handle[first_sheet_name]

    # Define assumed constants
    header_row_number = 1
    names_column_number = 1
    x_column_number = 2
    y_column_number = 3
    ordering_column_number = 5
    ordered_names_column_number = ordering_column_number + 1
    x_selected_column_number = ordered_names_column_number + 1
    y_selected_column_number = x_selected_column_number + 1

    # Extract the individual coordinates from file as a list of point objects
    extracted_points = []
    for row in range(header_row_number + 1, first_sheet.max_row + 1):
        x_coordinate = first_sheet.cell(row, x_column_number).value
        y_coordinate = first_sheet.cell(row, y_column_number).value
        extracted_points.append(geopy.point.Point(x_coordinate, y_coordinate))
    # Create a dictionary and corresponding index list mirroring/representing the extracted points
    available_points = {}
    available_points_indexes = []
    for index, point in enumerate(extracted_points):
        available_points[index] = point
        available_points_indexes.append(index)
    # Find two points and their indexes that are furthest apart - maximal distance between them
    n_of_available_points = len(available_points)
    maximal_distance = 0
    furthest_pair_indexes = ()
    for first_point_index in range(n_of_available_points):
        for second_point_index in range(first_point_index + 1, n_of_available_points):
            first_point = available_points[first_point_index]
            second_point = available_points[second_point_index]
            point_distance = geopy.distance.distance(first_point, second_point).km
            if point_distance > maximal_distance:
                maximal_distance = point_distance
                furthest_pair_indexes = (first_point_index, second_point_index)
    # Mark these two points as selected by removing them from the dictionary/list of available points/indexes
    # Manage points
    selected_points = {}
    selected_points[furthest_pair_indexes[0]] = available_points.pop(
        furthest_pair_indexes[0]
    )
    selected_points[furthest_pair_indexes[1]] = available_points.pop(
        furthest_pair_indexes[1]
    )
    # Manage indexes
    selected_points_indexes = list(furthest_pair_indexes)
    available_points_indexes.remove(furthest_pair_indexes[0])
    available_points_indexes.remove(furthest_pair_indexes[1])

    # Look for points until specified number has been found
    while len(selected_points) < how_many_points_to_find:
        # Define testing matrix where each available point will be in a separate row together with already selected points
        # All already selected points + 1 extra slot for 1 available point to test its "compatibility" wtith the others
        n_of_columns = len(selected_points) + 1
        n_of_rows = len(available_points)
        testing_matrix = []
        # Initialize first row with zeroes
        first_row = [0] * n_of_columns
        # Fill first row with already selected points (the extra slot at the end remains 0 at this point)
        for i in range(len(selected_points)):
            first_row[i] = selected_points[selected_points_indexes[i]]
        # Add the row to the "matrix"
        testing_matrix.extend(first_row)
        # Actually generate the matrix by copying the first row (selected points are the same in each row)
        testing_matrix = [testing_matrix.copy() for x in range(n_of_rows)]

        # Fill the last positions in each row of the matrix by one of the available points (the rows differ only in the last position)
        for i in range(len(available_points)):
            testing_matrix[i][n_of_columns - 1] = available_points[
                available_points_indexes[i]
            ]
        # Calculate distances of each two points in each row from each other and store it in the distance matrix (think of traveling in rows by the columns)
        distance_matrix = []
        for i in range(n_of_rows):
            distances_row = []
            for j in range(n_of_columns):
                # k is used so that already calculated pairs do not repeat, as that would be unnecessary
                for k in range(j + 1, n_of_columns):
                    distances_row.append(
                        geopy.distance.distance(
                            testing_matrix[i][j], testing_matrix[i][k]
                        ).km
                    )
            distance_matrix.append(distances_row)
        # Calculate standard deviation for each row - to find out how the additional new point affected the variance
        standard_deviations = []
        for row in distance_matrix:
            standard_deviations.append(numpy.std(row))
        # Find optimal index - optimal in a sense of representing minimal standard deviation
        optimal_point_index = numpy.argmin(standard_deviations)

        # Select the suitable points - remove from available points/their indexes, and append to the selected points/their indexes
        selected_points[
            available_points_indexes[optimal_point_index]
        ] = available_points.pop(available_points_indexes[optimal_point_index])
        selected_points_indexes.append(available_points_indexes[optimal_point_index])
        available_points_indexes.remove(available_points_indexes[optimal_point_index])
    # Extract the names of selected points
    list_of_selected_names = []
    for point_index in selected_points_indexes:
        list_of_selected_names.append(
            first_sheet.cell(point_index + 2, names_column_number).value
        )
    # Performing two separate distance sorts with different starting points
    list_of_selected_points = list(selected_points.values())
    # Sorted from south to north
    downmost_sorted_points, downmost_sorted_names = zip(
        *sorted(
            zip(list_of_selected_points, list_of_selected_names),
            key=lambda pair: pair[0].latitude,
        )
    )
    # Sorted from west to east
    leftmost_sorted_points, leftmost_sorted_names = zip(
        *sorted(
            zip(list_of_selected_points, list_of_selected_names),
            key=lambda pair: pair[0].longitude,
        )
    )

    # Transform tuples to lists to use pop()
    downmost_sorted_points = list(downmost_sorted_points)
    downmost_sorted_names = list(downmost_sorted_names)
    leftmost_sorted_points = list(leftmost_sorted_points)
    leftmost_sorted_names = list(leftmost_sorted_names)

    # Sorting of points in a way to assure the smallest distance between two adjacent points - optimal route
    distance_sorted_points_starting_down = []
    distance_sorted_points_starting_left = []

    # Variables to store names of sorted points
    distance_sorted_names_starting_down = []
    distance_sorted_names_starting_left = []

    # Pick the downmost and leftmost point in the south and west as the starting points
    distance_sorted_points_starting_down.append(downmost_sorted_points.pop(0))
    distance_sorted_points_starting_left.append(leftmost_sorted_points.pop(0))

    # Pick the downmost and leftmost name in the south and west as the starting names
    distance_sorted_names_starting_down.append(downmost_sorted_names.pop(0))
    distance_sorted_names_starting_left.append(leftmost_sorted_names.pop(0))

    # While some points are still left - until all points have been ordered
    # Calculate total distance to compare at the end
    total_distance_down = 0
    while downmost_sorted_points:
        smallest_distance = None
        smallest_distance_point_index = None
        # Calculate distance with regard to the last "selected" point
        first_point = distance_sorted_points_starting_down[-1]
        for index, second_point in enumerate(downmost_sorted_points):
            distance = geopy.distance.distance(first_point, second_point).km
            if not smallest_distance or distance < smallest_distance:
                smallest_distance = distance
                smallest_distance_point_index = index
                total_distance_down += distance
        distance_sorted_points_starting_down.append(
            downmost_sorted_points.pop(smallest_distance_point_index)
        )
        distance_sorted_names_starting_down.append(
            downmost_sorted_names.pop(smallest_distance_point_index)
        )
    # While some points are still left - until all points have been ordered
    # Calculate total distance to compare at the end
    total_distance_left = 0
    while leftmost_sorted_points:
        smallest_distance = None
        smallest_distance_point_index = None
        # Calculate distance with regard to the last "selected" point
        first_point = distance_sorted_points_starting_left[-1]
        for index, second_point in enumerate(leftmost_sorted_points):
            distance = geopy.distance.distance(first_point, second_point).km
            if not smallest_distance or distance < smallest_distance:
                smallest_distance = distance
                smallest_distance_point_index = index
                total_distance_left += distance
        distance_sorted_points_starting_left.append(
            leftmost_sorted_points.pop(smallest_distance_point_index)
        )
        distance_sorted_names_starting_left.append(
            leftmost_sorted_names.pop(smallest_distance_point_index)
        )
    # Select the path with the smaller total distance as the one to be exported
    if total_distance_down < total_distance_left:
        points_to_write_to_excel = distance_sorted_points_starting_down
        names_to_write_to_excel = distance_sorted_names_starting_down
    else:
        points_to_write_to_excel = distance_sorted_points_starting_left
        names_to_write_to_excel = distance_sorted_names_starting_left
    # Fill the cells representing the selected points with red colour
    for row in selected_points_indexes:
        for column in range(1, first_sheet.max_column + 1):
            first_sheet.cell(row + 2, column).fill = openpyxl.styles.PatternFill(
                fgColor="FFCCCC", fill_type="solid"
            )
            first_sheet.cell(row + 2, column).font = openpyxl.styles.Font(
                color="9C0006"
            )
    # Define cell border
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add headers for selected points and format them
    first_sheet.cell(1, ordering_column_number).value = "Number"
    first_sheet.cell(1, ordering_column_number).font = openpyxl.styles.Font(bold=True)
    first_sheet.cell(1, ordering_column_number).border = thin_border
    first_sheet.cell(1, ordered_names_column_number).value = "Description"
    first_sheet.cell(1, ordered_names_column_number).font = openpyxl.styles.Font(
        bold=True
    )
    first_sheet.cell(1, ordered_names_column_number).border = thin_border
    first_sheet.cell(1, x_selected_column_number).value = "Latitude [°]"
    first_sheet.cell(1, x_selected_column_number).font = openpyxl.styles.Font(bold=True)
    first_sheet.cell(1, x_selected_column_number).border = thin_border
    first_sheet.cell(1, y_selected_column_number).value = "Longitude [°]"
    first_sheet.cell(1, y_selected_column_number).font = openpyxl.styles.Font(bold=True)
    first_sheet.cell(1, y_selected_column_number).border = thin_border

    # Write the selected points in separate columns into the excel file and format them
    n_of_selected_points = len(selected_points)
    for row in range(n_of_selected_points):
        first_sheet.cell(row + 2, ordering_column_number).value = row + 1
        first_sheet.cell(row + 2, ordering_column_number).number_format = "0"
        first_sheet.cell(row + 2, ordering_column_number).border = thin_border
        first_sheet.cell(
            row + 2, ordered_names_column_number
        ).value = names_to_write_to_excel[row]
        first_sheet.cell(row + 2, ordered_names_column_number).border = thin_border
        first_sheet.cell(
            row + 2, x_selected_column_number
        ).value = points_to_write_to_excel[row].latitude
        first_sheet.cell(row + 2, x_selected_column_number).number_format = "0.00000"
        first_sheet.cell(row + 2, x_selected_column_number).border = thin_border
        first_sheet.cell(
            row + 2, y_selected_column_number
        ).value = points_to_write_to_excel[row].longitude
        first_sheet.cell(row + 2, y_selected_column_number).number_format = "0.00000"
        first_sheet.cell(row + 2, y_selected_column_number).border = thin_border
    # Note: This step is valid only if the Excel decimal delimiter in your language is colon
    # Replace dot delimiters with colon in the original data range (if present)
    for row in range(header_row_number + 1, first_sheet.max_row + 1):
        for column in range(x_column_number, y_column_number + 1):
            cell_string_value = str(first_sheet.cell(row, column).value)
            if "." in cell_string_value:
                first_sheet.cell(row, column).value = float(
                    first_sheet.cell(row, column).value
                )
            first_sheet.cell(row, column).number_format = "0.00000"
    # Apply border to the original data range
    for row in range(header_row_number, first_sheet.max_row + 1):
        for column in range(1, y_column_number + 1):
            first_sheet.cell(row, column).border = thin_border
    # In case the original data range was not contained in a table, contain it in a table
    if not first_sheet.tables:
        main_data_table = Table(
            displayName="Main_Data_Table", ref="A1:C" + str(first_sheet.max_row)
        )
        first_sheet.add_table(main_data_table)
    # Contain the selected data in a table
    selected_data_table = Table(
        displayName="Selected_Data_Table",
        ref="E1:H" + str(n_of_selected_points + header_row_number),
    )
    first_sheet.add_table(selected_data_table)

    # Modify the width of new columns to better accomodate their contents
    # For unknown reason bestFit does not work as intended
    # first_sheet.column_dimensions['E'].bestFit = True
    # first_sheet.column_dimensions['F'].bestFit = True
    # first_sheet.column_dimensions['G'].bestFit = True
    # first_sheet.column_dimensions['H'].bestFit = True

    first_sheet.column_dimensions["E"].width = 10
    first_sheet.column_dimensions["F"].width = 20
    first_sheet.column_dimensions["G"].width = 20
    first_sheet.column_dimensions["H"].width = 20

    # Make sure that the header in the original range is also bold
    first_sheet.cell(
        header_row_number, names_column_number
    ).font = openpyxl.styles.Font(bold=True)
    first_sheet.cell(header_row_number, x_column_number).font = openpyxl.styles.Font(
        bold=True
    )
    first_sheet.cell(header_row_number, y_column_number).font = openpyxl.styles.Font(
        bold=True
    )

    # Save the changes performed on the file
    excel_workbook_handle.save(path_to_copied_file)

    # Find approximate middle point for the map (teoretically third selected point):
    approximate_middle_point = (
        selected_points[selected_points_indexes[2]]
        if len(selected_points_indexes) > 2
        else selected_points[selected_points_indexes[0]]
    )
    # Congifure google map through API key and set middle point with initial zoom
    apikey = "abcdefghijklmnopqrstuvwxyz0123456789"  # (your API key here)
    google_map = gmplot.GoogleMapPlotter(
        approximate_middle_point.latitude,
        approximate_middle_point.longitude,
        14,
        apikey=apikey,
    )

    # Extract original latitudes and longitudes and plot them with full circles
    latitudes = []
    longitudes = []
    for point in extracted_points:
        latitudes.append(point.latitude)
        longitudes.append(point.longitude)
    google_map.scatter(
        latitudes, longitudes, size=15, color="magenta", marker=False, fa=1
    )

    # Add red numbered markers for actually selected points
    for index, point in enumerate(points_to_write_to_excel):
        google_map.marker(
            point.latitude,
            point.longitude,
            color="red",
            label=str(index + 1),
            title="Latidude: {:.5f} Longitude: {:.5f}".format(
                point.latitude, point.longitude
            ),
        )
    # Create path to save map
    map_name = chosen_file_name_and_extension[0] + "_MAP.html"
    path_to_map = os.path.join(chosen_file_folder_path, map_name)

    # "Draw" the map
    google_map.draw(path_to_map)

    # Find chrome or edge browser
    path_to_browser = None
    expected_path_1 = "C://Program Files (x86)//Google//Chrome//Application//chrome.exe"
    expected_path_2 = "C://Program Files//Google//Chrome//Application//chrome.exe"
    expected_path_3 = (
        "C://Program Files (x86)//Microsoft//Edge//Application//msedge.exe"
    )
    expected_path_4 = "C://Program Files//Microsoft//Edge//Application//msedge.exe"

    # Test which browser location is valid
    if os.path.exists(expected_path_1):
        path_to_browser = expected_path_1 + " %s"
    elif os.path.exists(expected_path_2):
        path_to_browser = expected_path_2 + " %s"
    elif os.path.exists(expected_path_3):
        path_to_browser = expected_path_3 + " %s"
    elif os.path.exists(expected_path_4):
        path_to_browser = expected_path_4 + " %s"
    # Open the map in browser, ideally chrome
    webbrowser.get(path_to_browser).open_new(path_to_map)

    # Return the name of the new file
    return copied_file_name
